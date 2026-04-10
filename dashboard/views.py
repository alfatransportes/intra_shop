import json
import re
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

import qrcode
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.generic import (CreateView, DeleteView, DetailView, ListView,
                                  TemplateView, UpdateView, View)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from dashboard.services.importador_produtos import importar_produtos
from website.models import (Carrinho, FormaPagamento, NivelAvaria, Produto,
                            ProdutoImagem, ProdutoVariacao,
                            RegraParcelamentoVale, Tipo, Unidade, Venda,
                            VendaItem)

from .forms import (FormaPagamentoForm, NivelAvariaForm, ProdutoForm,
                    ProdutoImagemForm, ProdutoImagemFormSet, ProdutoImportForm,
                    ProdutoVariacaoFormSet, RegraParcelamentoValeForm,
                    TipoForm, UnidadeForm, VendaForm, VendaItemFormSet,
                    VendaStatusForm)
from .mixins import DashboardPermissionMixin
from .utils import enviar_email_status_venda_cliente


class PainelControleView(DashboardPermissionMixin, TemplateView):
    template_name = "dashboard/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pedidos = Venda.objects.all()
        vendas_pendentes_qs = pedidos.filter(status=Venda.Status.PENDENTE)
        vendas_aprovadas_qs = pedidos.filter(status=Venda.Status.APROVADA)
        vendas_canceladas_qs = pedidos.filter(status=Venda.Status.CANCELADA)

        context["total_produtos"] = Produto.objects.count()
        context["total_pedidos"] = pedidos.count()
        context["vendas_pendentes"] = vendas_pendentes_qs.count()
        context["vendas_aprovadas"] = vendas_aprovadas_qs.count()
        context["vendas_canceladas"] = vendas_canceladas_qs.count()
        context["carrinhos_fechados"] = Carrinho.objects.filter(status=Carrinho.Status.FECHADO).count()

        context["valor_total_geral"] = (
            pedidos.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        )

        context["valor_total_aprovado"] = (
            vendas_aprovadas_qs.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        )

        context["valor_total_pendente"] = (
            vendas_pendentes_qs.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        )

        context["valor_total_cancelado"] = (
            vendas_canceladas_qs.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        )

        valor_mercadoria = Produto.objects.aggregate(
            total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("quantidade") * F("valor_venda"),
                        output_field=DecimalField(max_digits=14, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            )
        )["total"]

        context["valor_total_mercadoria"] = valor_mercadoria or Decimal("0.00")

        context["produtos_sem_estoque"] = Produto.objects.filter(quantidade=0).count()
        context["produtos_com_estoque"] = Produto.objects.filter(quantidade__gt=0).count()

        context["ultimas_vendas"] = (
            pedidos.select_related("usuario", "forma_pagamento")
            .order_by("-criado_em")[:5]
        )

        return context


class DashboardBaseListView(DashboardPermissionMixin, ListView):
    template_name = "dashboard/crud/list.html"
    context_object_name = "items"
    paginate_by = 20
    ordering = ["nome"]

    page_title = ""
    page_subtitle = ""
    create_url_name = ""
    empty_message = "Nenhum item cadastrado."

    columns = []  # ex: [{"label": "Nome", "field": "nome"}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.page_title
        context["page_subtitle"] = self.page_subtitle
        context["create_url_name"] = self.create_url_name
        context["empty_message"] = self.empty_message
        context["columns"] = self.columns
        return context


class DashboardBaseFormViewMixin(DashboardPermissionMixin):
    template_name = "dashboard/crud/form.html"
    page_title_create = ""
    page_title_update = ""
    page_subtitle = ""
    cancel_url_name = ""
    success_message = ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = (
            self.page_title_update if getattr(self, "object", None) else self.page_title_create
        )
        context["page_subtitle"] = self.page_subtitle
        context["cancel_url_name"] = self.cancel_url_name
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.success_message:
            messages.success(self.request, self.success_message)
        return response


class DashboardBaseCreateView(DashboardBaseFormViewMixin, CreateView):
    pass


class DashboardBaseUpdateView(DashboardBaseFormViewMixin, UpdateView):
    pass


class DashboardBaseDeleteView(DashboardPermissionMixin, DeleteView):
    template_name = "dashboard/confirm_delete.html"
    success_message = ""

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, self.success_message)
        return super().delete(request, *args, **kwargs)


# -------------------------
# PRODUTOS
# -------------------------

class ProdutoListView(DashboardPermissionMixin, ListView):
    model = Produto
    template_name = "dashboard/produtos/list.html"
    context_object_name = "produtos"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            Produto.objects
            .select_related("unidade_prod", "tipo_prod", "nivel_ava_prod")
            .prefetch_related("imagens")
            .order_by("nome")
        )

        busca = self.request.GET.get("q")
        tipo = self.request.GET.get("tipo")

        if busca:
            queryset = queryset.filter(
                Q(nome__icontains=busca) |
                Q(num_controle__icontains=busca)
            )

        if tipo:
            queryset = queryset.filter(tipo_prod_id=tipo)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tipos"] = Tipo.objects.filter(ativo=True).order_by("nome")
        context["q"] = self.request.GET.get("q", "")
        context["tipo_atual"] = self.request.GET.get("tipo", "")
        return context


class ProdutoManageView(DashboardPermissionMixin, TemplateView):
    template_name = "dashboard/produtos/form.html"

    def dispatch(self, request, *args, **kwargs):
        self.object = None
        pk = kwargs.get("pk")
        if pk:
            self.object = Produto.objects.filter(pk=pk).first()
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, data=None, files=None):
        return ProdutoForm(data=data, files=files, instance=self.object)

    def get_next_step_url(self, produto):
        if produto.usa_variacoes and not produto.variacoes.filter(ativo=True, quantidade__gt=0).exists():
            return f"{reverse('dashboard_produto_update', kwargs={'pk': produto.pk})}?step=variacoes"
        if not produto.imagens.exists():
            return f"{reverse('dashboard_produto_update', kwargs={'pk': produto.pk})}?step=imagens"
        return reverse("dashboard_produto_update", kwargs={"pk": produto.pk})

    def get_formset(self, data=None, files=None):
        if not self.object:
            return ProdutoImagemFormSet(prefix="imagens")
        return ProdutoImagemFormSet(
            data=data,
            files=files,
            instance=self.object,
            prefix="imagens",
        )
    
    def get_variacao_formset(self, data=None, files=None):
        if not self.object:
            return ProdutoVariacaoFormSet(prefix="variacoes")
        return ProdutoVariacaoFormSet(
            data=data,
            files=files,
            instance=self.object,
            prefix="variacoes",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = kwargs.get("form") or self.get_form()
        open_images_modal = kwargs.get("open_images_modal", False)
        open_variacoes_modal = kwargs.get("open_variacoes_modal", False)
        step = self.request.GET.get("step")

        if self.object and step == "variacoes":
            open_variacoes_modal = True
        if self.object and step == "imagens":
            open_images_modal = True

        context["object"] = self.object
        context["form"] = form
        context["formset"] = kwargs.get("formset") or self.get_formset()
        context["open_images_modal"] = open_images_modal
        context["variacao_formset"] = kwargs.get("variacao_formset") or self.get_variacao_formset()
        context["open_variacoes_modal"] = open_variacoes_modal
        context["produto_pode_ativar"] = False
        context["produto_motivo_bloqueio"] = ""
        return context

    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())

    def post(self, request, *args, **kwargs):
        form_type = request.POST.get("form_type", "produto")

        # Cadastro novo: só salva dados do produto
        if not self.object:
            form = self.get_form(data=request.POST, files=request.FILES)

            if form.is_valid():
                self.object = form.save()
                messages.success(request, "Produto cadastrado com sucesso.")
                return redirect(self.get_next_step_url(self.object))

            return self.render_to_response(
                self.get_context_data(form=form, formset=self.get_formset())
            )

        # Edição: submit separado por tipo
        if form_type == "imagens":
            return self.handle_images_form(request)
        
        if form_type == "variacoes":
            return self.handle_variacoes_form(request)

        return self.handle_product_form(request)

    def handle_product_form(self, request):
        form = self.get_form(data=request.POST, files=request.FILES)
        formset = self.get_formset()

        if form.is_valid():
            self.object = form.save()

            # Segurança extra: se ficou sem imagens, não deixa ativo
            if self.object.ativo and not self.object.imagens.exists():
                self.object.ativo = False
                self.object.save(update_fields=["ativo"])

            messages.success(request, "Dados do produto salvos com sucesso.")
            return redirect(self.get_next_step_url(self.object))

        messages.error(request, "Corrija os erros no formulário do produto.")
        return self.render_to_response(
            self.get_context_data(form=form, formset=formset)
        )
    

    def handle_variacoes_form(self, request):
        form = self.get_form()
        formset = self.get_formset()
        variacao_formset = self.get_variacao_formset(data=request.POST, files=request.FILES)

        if variacao_formset.is_valid():
            with transaction.atomic():
                variacao_formset.save()

                if self.object.usa_variacoes and not self.object.tem_variacoes_ativas_com_estoque and self.object.ativo:
                    self.object.ativo = False
                    self.object.save(update_fields=["ativo"])

            messages.success(request, "Variações salvas com sucesso.")
            return redirect(self.get_next_step_url(self.object))

        messages.error(request, "Corrija os erros no formulário de variações.")
        return self.render_to_response(
            self.get_context_data(
                form=form,
                formset=formset,
                variacao_formset=variacao_formset,
                open_variacoes_modal=True,
            )
        )

    def handle_images_form(self, request):
        form = self.get_form()
        formset = self.get_formset(data=request.POST, files=request.FILES)

        if formset.is_valid():
            with transaction.atomic():
                formset.save()

                principais = [
                    f for f in formset.forms
                    if getattr(f, "cleaned_data", None)
                    and not f.cleaned_data.get("DELETE", False)
                    and f.cleaned_data.get("principal")
                ]

                if len(principais) == 1:
                    imagem_principal = principais[0].instance
                    self.object.imagens.exclude(pk=imagem_principal.pk).update(principal=False)

                # Se não houver nenhuma imagem restante, força produto inativo
                if not self.object.imagens.exists() and self.object.ativo:
                    self.object.ativo = False
                    self.object.save(update_fields=["ativo"])

            messages.success(request, "Imagens salvas com sucesso.")
            return redirect(self.get_next_step_url(self.object))

        messages.error(request, "Corrija os erros no formulário de imagens.")
        return self.render_to_response(
            self.get_context_data(
                form=form,
                formset=formset,
                open_images_modal=True,
            )
        )


class ProdutoDeleteView(DeleteView):
    model = Produto
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("dashboard_produto_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        possui_venda = self.object.itens_venda.exists()
        possui_carrinho = self.object.itens_carrinho.exists()

        if possui_venda or possui_carrinho:
            self.object.ativo = False
            self.object.save(update_fields=["ativo"])
            messages.warning(
                request,
                "O produto já está vinculado a registros e não pode ser excluído. "
                "Ele foi apenas desativado."
            )
        else:
            self.object.delete()
            messages.success(request, "Produto excluído com sucesso.")

        return redirect(self.success_url)


class ProdutoExportXlsxView(DashboardPermissionMixin, View):

    def get_queryset(self):
        queryset = (
            Produto.objects
            .select_related("unidade_prod", "tipo_prod", "nivel_ava_prod")
            .prefetch_related("imagens", "variacoes")
            .order_by("nome", "id")
        )

        busca = (self.request.GET.get("q") or "").strip()
        tipo = (self.request.GET.get("tipo") or "").strip()

        if busca:
            queryset = queryset.filter(nome__icontains=busca)

        if tipo.isdigit():
            queryset = queryset.filter(tipo_prod_id=int(tipo))

        return queryset

    def _slug_filename_part(self, value, default="todos"):
        value = (value or "").strip()
        if not value:
            return default

        value = re.sub(r"[^a-zA-Z0-9_-]+", "_", value)
        value = re.sub(r"_+", "_", value).strip("_")
        return value or default

    def _formatar_variacoes(self, produto):
        if not produto.usa_variacoes:
            return ""

        partes = []
        for variacao in produto.variacoes.filter(ativo=True).order_by(
            "categoria", "genero", "faixa_etaria", "tamanho", "cor", "id"
        ):
            descricao = []

            if variacao.categoria:
                descricao.append(variacao.get_categoria_display())
            if variacao.genero:
                descricao.append(variacao.get_genero_display())
            if variacao.faixa_etaria:
                descricao.append(variacao.get_faixa_etaria_display())
            if variacao.tamanho:
                descricao.append(f"Tamanho {variacao.tamanho}")
            if variacao.cor:
                descricao.append(variacao.cor)

            texto = " / ".join(descricao) if descricao else f"Variação #{variacao.id}"
            texto += f" (estoque: {int(variacao.quantidade or 0)})"
            partes.append(texto)

        return "\n".join(partes)

    def get(self, request, *args, **kwargs):
        produtos = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        headers = [
            "ID",
            "Nome",
            "Tipo",
            "Unidade",
            "Nível Avaria",
            "Usa Variações",
            "Quantidade Base",
            "Estoque Disponível",
            "Valor Nota",
            "Desconto %",
            "Valor Venda",
            "Ativo",
            "Variações",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for produto in produtos:
            ws.append([
                produto.id,
                produto.nome,
                produto.tipo_prod.nome if produto.tipo_prod else "",
                produto.unidade_prod.nome if produto.unidade_prod else "",
                produto.nivel_ava_prod.nome if produto.nivel_ava_prod else "",
                "Sim" if produto.usa_variacoes else "Não",
                int(produto.quantidade or 0),
                int(produto.estoque_disponivel or 0),
                float(produto.valor_nota or Decimal("0.00")),
                float(produto.porcen_desconto or Decimal("0.00")),
                float(produto.valor_venda or Decimal("0.00")),
                "Sim" if produto.ativo else "Não",
                self._formatar_variacoes(produto),
            ])

        widths = {
            "A": 10,
            "B": 40,
            "C": 25,
            "D": 20,
            "E": 25,
            "F": 16,
            "G": 18,
            "H": 20,
            "I": 15,
            "J": 15,
            "K": 18,
            "L": 10,
            "M": 70,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in range(2, ws.max_row + 1):
            ws[f"I{row}"].number_format = 'R$ #,##0.00'
            ws[f"J{row}"].number_format = '0.00'
            ws[f"K{row}"].number_format = 'R$ #,##0.00'

            ws[f"B{row}"].alignment = Alignment(vertical="top")
            ws[f"C{row}"].alignment = Alignment(vertical="top")
            ws[f"D{row}"].alignment = Alignment(vertical="top")
            ws[f"E{row}"].alignment = Alignment(vertical="top")
            ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"M{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"

        busca = self._slug_filename_part(request.GET.get("q"), "todos")
        tipo = self._slug_filename_part(request.GET.get("tipo"), "todos")

        filename = f"produtos_{busca}_{tipo}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


@xframe_options_exempt
def produto_qrcode_pdf(request, pk):
    produto = get_object_or_404(Produto, pk=pk)

    url_detalhe = request.build_absolute_uri(
        reverse("detalhe_produto", args=[produto.pk])
    )

    # Gera o QRCode em memória
    qr = qrcode.QRCode(
        version=1,
        box_size=10,
        border=2,
    )
    qr.add_data(url_detalhe)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    # PDF em memória
    pdf_buffer = BytesIO()

    # Tamanho da etiqueta / página
    page_width = 80 * mm
    page_height = 100 * mm

    c = canvas.Canvas(pdf_buffer, pagesize=(page_width, page_height))
    c.setTitle(f"QR Code - {produto.nome}")
    margem = 8 * mm
    y_top = page_height - margem

    # Título / nome do produto
    nome = produto.nome or "Produto"
    fonte_nome = "Helvetica-Bold"
    tamanho_nome = 11

    max_text_width = page_width - (2 * margem)

    # quebra simples em até 2 linhas
    palavras = nome.split()
    linhas = []
    linha_atual = ""

    for palavra in palavras:
        teste = f"{linha_atual} {palavra}".strip()
        if stringWidth(teste, fonte_nome, tamanho_nome) <= max_text_width:
            linha_atual = teste
        else:
            if linha_atual:
                linhas.append(linha_atual)
            linha_atual = palavra

    if linha_atual:
        linhas.append(linha_atual)

    linhas = linhas[:2]
    if len(linhas) == 2 and len(palavras) > 0:
        # se sobrou conteúdo, corta discretamente
        while stringWidth(linhas[-1] + "...", fonte_nome, tamanho_nome) > max_text_width and len(linhas[-1]) > 1:
            linhas[-1] = linhas[-1][:-1]
        if len(" ".join(palavras)) > len(" ".join(linhas)):
            linhas[-1] += "..."

    c.setFont(fonte_nome, tamanho_nome)
    y = y_top

    for linha in linhas:
        text_width = stringWidth(linha, fonte_nome, tamanho_nome)
        x = (page_width - text_width) / 2
        c.drawString(x, y, linha)
        y -= 5.5 * mm

    y -= 2 * mm

    # QR centralizado
    qr_size = 48 * mm
    qr_x = (page_width - qr_size) / 2
    qr_y = y - qr_size

    c.drawImage(
        ImageReader(qr_buffer),
        qr_x,
        qr_y,
        width=qr_size,
        height=qr_size,
        preserveAspectRatio=True,
        mask='auto',
    )

    # legenda pequena
    legenda = "Escaneie para ver detalhes"
    c.setFont("Helvetica", 8)
    legenda_width = stringWidth(legenda, "Helvetica", 8)
    c.drawString((page_width - legenda_width) / 2, qr_y - 6 * mm, legenda)

    c.showPage()
    c.save()

    pdf_buffer.seek(0)

    filename = f"qrcode-produto-{produto.pk}.pdf"
    safe_filename = quote(filename)

    response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{safe_filename}"'
    return response



class ProdutoImportView(DashboardPermissionMixin, TemplateView):
    template_name = "dashboard/produtos/importar.html"

    def get_form(self, data=None, files=None):
        return ProdutoImportForm(data=data, files=files)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = kwargs.get("form") or self.get_form()
        context["resultado"] = kwargs.get("resultado")
        return context

    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())

    def post(self, request, *args, **kwargs):
        form = self.get_form(data=request.POST, files=request.FILES)

        if not form.is_valid():
            messages.error(request, "Selecione um arquivo válido para importação.")
            return self.render_to_response(self.get_context_data(form=form))

        arquivo = form.cleaned_data["arquivo"]

        try:
            resultado = importar_produtos(arquivo)

            messages.success(
                request,
                (
                    f"Importação concluída. "
                    f"Produtos criados: {resultado['criados']} | "
                    f"Produtos atualizados: {resultado['atualizados']} | "
                    f"Variações criadas: {resultado['variacoes_criadas']} | "
                    f"Variações atualizadas: {resultado['variacoes_atualizadas']} | "
                    f"Erros: {len(resultado['erros'])}"
                ),
            )

            if resultado["erros"]:
                messages.warning(
                    request,
                    "Algumas linhas não puderam ser importadas ou alguns produtos não puderam ser publicados. Veja o relatório abaixo."
                )

            return self.render_to_response(
                self.get_context_data(
                    form=self.get_form(),
                    resultado=resultado,
                )
            )

        except Exception as e:
            messages.error(request, f"Erro ao processar a planilha: {e}")
            return self.render_to_response(self.get_context_data(form=form))



class ProdutoImportTemplateDownloadView(DashboardPermissionMixin, View):
    filename = "modelo_importacao_produtos.xlsx"

    def get(self, request, *args, **kwargs):
        wb = Workbook()

        ws_import_produtos = wb.active
        ws_import_produtos.title = "IMPORTACAO_PRODUTOS"

        ws_import_variacoes = wb.create_sheet("IMPORTACAO_VARIACOES")
        ws_listas = wb.create_sheet("LISTAS_SISTEMA")
        ws_instrucoes = wb.create_sheet("INSTRUCOES")

        contadores = self._build_listas_sheet(ws_listas)
        self._create_named_ranges(wb, contadores)
        self._build_import_produtos_sheet(ws_import_produtos)
        self._build_import_variacoes_sheet(ws_import_variacoes)
        self._build_instrucoes_sheet(ws_instrucoes)
        self._apply_validations(ws_import_produtos, ws_import_variacoes)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}"'
        return response

    def _style_header(self, ws, color):
        fill = PatternFill("solid", fgColor=color)
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align

        ws.freeze_panes = "A2"

    def _build_import_produtos_sheet(self, ws):
        headers = [
            "num_controle",
            "nome",
            "unidade_prod",
            "tipo_prod",
            "nivel_ava_prod",
            "usa_variacoes",
            "quantidade",
            "maximo_por_usuario",
            "valor_nota",
            "porcen_desconto",
            "descricao",
            "ativo",
        ]
        ws.append(headers)
        self._style_header(ws, "0D6EFD")

        widths = {
            1: 22,
            2: 32,
            3: 30,
            4: 24,
            5: 24,
            6: 16,
            7: 14,
            8: 22,
            9: 16,
            10: 18,
            11: 45,
            12: 12,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.append([
            "SKU-001",
            "Notebook Dell",
            "",
            "",
            "",
            "Não",
            5,
            1,
            3500.00,
            10.00,
            "Notebook corporativo",
            "Não",
        ])

        ws.append([
            "SKU-002",
            "Tênis Esportivo",
            "",
            "",
            "",
            "Sim",
            0,
            1,
            499.90,
            15.00,
            "Produto com grade de tamanhos",
            "Não",
        ])

    def _build_import_variacoes_sheet(self, ws):
        headers = [
            "produto_ref",
            "categoria",
            "genero",
            "faixa_etaria",
            "tamanho",
            "cor",
            "quantidade",
            "ativo",
        ]
        ws.append(headers)
        self._style_header(ws, "6F42C1")

        widths = {
            1: 22,
            2: 20,
            3: 20,
            4: 20,
            5: 14,
            6: 18,
            7: 14,
            8: 12,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.append([
            "SKU-002",
            "",
            "",
            "",
            "39",
            "Preto",
            3,
            "Sim",
        ])
        ws.append([
            "SKU-002",
            "",
            "",
            "",
            "40",
            "Preto",
            4,
            "Sim",
        ])

    def _build_listas_sheet(self, ws):
        headers = [
            "UNIDADES",
            "TIPOS",
            "NIVEIS_AVARIA",
            "BOOLEANOS",
            "CATEGORIAS_VARIACAO",
            "GENEROS_VARIACAO",
            "FAIXAS_ETARIAS_VARIACAO",
        ]
        ws.append(headers)
        self._style_header(ws, "198754")

        unidades = list(
            Unidade.objects.order_by("codigo", "nome").values_list("codigo", "nome")
        )
        tipos = list(
            Tipo.objects.filter(ativo=True).order_by("nome").values_list("nome", flat=True)
        )
        niveis = list(
            NivelAvaria.objects.order_by("nome").values_list("nome", flat=True)
        )
        booleanos = ["Sim", "Não"]

        categoria_choices = ProdutoVariacao._meta.get_field("categoria").choices or []
        genero_choices = ProdutoVariacao._meta.get_field("genero").choices or []
        faixa_choices = ProdutoVariacao._meta.get_field("faixa_etaria").choices or []

        max_rows = max(
            len(unidades),
            len(tipos),
            len(niveis),
            len(booleanos),
            len(categoria_choices),
            len(genero_choices),
            len(faixa_choices),
            1,
        )

        for idx in range(max_rows):
            unidade_val = ""
            if idx < len(unidades):
                codigo, nome = unidades[idx]
                unidade_val = f"{codigo} - {nome}" if codigo is not None else nome

            tipo_val = tipos[idx] if idx < len(tipos) else ""
            nivel_val = niveis[idx] if idx < len(niveis) else ""
            bool_val = booleanos[idx] if idx < len(booleanos) else ""
            categoria_val = categoria_choices[idx][1] if idx < len(categoria_choices) else ""
            genero_val = genero_choices[idx][1] if idx < len(genero_choices) else ""
            faixa_val = faixa_choices[idx][1] if idx < len(faixa_choices) else ""

            ws.append([
                unidade_val,
                tipo_val,
                nivel_val,
                bool_val,
                categoria_val,
                genero_val,
                faixa_val,
            ])

        return {
            "unidades": len(unidades),
            "tipos": len(tipos),
            "niveis": len(niveis),
            "booleanos": len(booleanos),
            "categorias_variacao": len(categoria_choices),
            "generos_variacao": len(genero_choices),
            "faixas_etarias_variacao": len(faixa_choices),
        }

    def _create_named_ranges(self, wb, contadores):
        ranges = {
            "LISTA_UNIDADES": ("LISTAS_SISTEMA", "A", contadores["unidades"]),
            "LISTA_TIPOS": ("LISTAS_SISTEMA", "B", contadores["tipos"]),
            "LISTA_NIVEIS_AVARIA": ("LISTAS_SISTEMA", "C", contadores["niveis"]),
            "LISTA_BOOLEANOS": ("LISTAS_SISTEMA", "D", contadores["booleanos"]),
            "LISTA_CATEGORIAS_VARIACAO": ("LISTAS_SISTEMA", "E", contadores["categorias_variacao"]),
            "LISTA_GENEROS_VARIACAO": ("LISTAS_SISTEMA", "F", contadores["generos_variacao"]),
            "LISTA_FAIXAS_ETARIAS_VARIACAO": ("LISTAS_SISTEMA", "G", contadores["faixas_etarias_variacao"]),
        }

        for nome, (aba, coluna, total) in ranges.items():
            if total <= 0:
                continue
            ref = f"'{aba}'!${coluna}$2:${coluna}${total + 1}"
            wb.defined_names.add(DefinedName(nome, attr_text=ref))

    def _apply_validations(self, ws_produtos, ws_variacoes):
        limite = 1000

        dv_unidades = DataValidation(type="list", formula1="=LISTA_UNIDADES", allow_blank=True)
        dv_tipos = DataValidation(type="list", formula1="=LISTA_TIPOS", allow_blank=True)
        dv_niveis = DataValidation(type="list", formula1="=LISTA_NIVEIS_AVARIA", allow_blank=True)
        dv_bool = DataValidation(type="list", formula1="=LISTA_BOOLEANOS", allow_blank=True)

        dv_cat = DataValidation(type="list", formula1="=LISTA_CATEGORIAS_VARIACAO", allow_blank=True)
        dv_gen = DataValidation(type="list", formula1="=LISTA_GENEROS_VARIACAO", allow_blank=True)
        dv_faixa = DataValidation(type="list", formula1="=LISTA_FAIXAS_ETARIAS_VARIACAO", allow_blank=True)

        ws_produtos.add_data_validation(dv_unidades)
        ws_produtos.add_data_validation(dv_tipos)
        ws_produtos.add_data_validation(dv_niveis)
        ws_produtos.add_data_validation(dv_bool)

        ws_variacoes.add_data_validation(dv_cat)
        ws_variacoes.add_data_validation(dv_gen)
        ws_variacoes.add_data_validation(dv_faixa)
        ws_variacoes.add_data_validation(dv_bool)

        dv_unidades.add(f"C2:C{limite}")
        dv_tipos.add(f"D2:D{limite}")
        dv_niveis.add(f"E2:E{limite}")
        dv_bool.add(f"F2:F{limite}")
        dv_bool.add(f"L2:L{limite}")

        dv_cat.add(f"B2:B{limite}")
        dv_gen.add(f"C2:C{limite}")
        dv_faixa.add(f"D2:D{limite}")
        dv_bool.add(f"H2:H{limite}")

    def _build_instrucoes_sheet(self, ws):
        linhas = [
            ["COMO USAR O IMPORTADOR"],
            [""],
            ["1. Preencha a aba IMPORTACAO_PRODUTOS."],
            ["2. Para produtos com variações, marque usa_variacoes = Sim."],
            ["3. Para produtos com variações, deixe quantidade base = 0 e use a aba IMPORTACAO_VARIACOES."],
            ["4. Em produto_ref, use preferencialmente o num_controle. Se não houver, use o nome exato do produto."],
            ["5. O sistema recalcula automaticamente o valor de venda."],
            ["6. Produto pode continuar em rascunho se não cumprir as regras de ativação (ex.: sem imagem)."],
        ]

        for linha in linhas:
            ws.append(linha)

        ws.column_dimensions["A"].width = 110
        ws["A1"].font = Font(bold=True)

# -------------------------
# TIPOS
# -------------------------
class TipoListView(DashboardBaseListView):
    model = Tipo
    ordering = ["nome"]
    page_title = "Tipos"
    page_subtitle = "Gerencie os tipos cadastrados."
    create_url_name = "dashboard_tipo_create"
    empty_message = "Nenhum tipo cadastrado."
    columns = [
        {"label": "Nome", "field": "nome"},
    ]


class TipoCreateView(DashboardBaseCreateView):
    model = Tipo
    form_class = TipoForm
    success_url = reverse_lazy("dashboard_tipo_list")
    page_title_create = "Novo tipo"
    page_title_update = "Editar tipo"
    page_subtitle = "Preencha os dados do tipo."
    cancel_url_name = "dashboard_tipo_list"
    success_message = "Tipo cadastrado com sucesso."


class TipoUpdateView(DashboardBaseUpdateView):
    model = Tipo
    form_class = TipoForm
    success_url = reverse_lazy("dashboard_tipo_list")
    page_title_create = "Novo tipo"
    page_title_update = "Editar tipo"
    page_subtitle = "Preencha os dados do tipo."
    cancel_url_name = "dashboard_tipo_list"
    success_message = "Tipo atualizado com sucesso."


class TipoDeleteView(DashboardBaseDeleteView):
    model = Tipo
    success_url = reverse_lazy("dashboard_tipo_list")
    success_message = "Tipo excluído com sucesso."

# -------------------------
# UNIDADES
# -------------------------
class UnidadeListView(DashboardBaseListView):
    model = Unidade
    ordering = ["nome"]
    page_title = "Unidades"
    page_subtitle = "Gerencie as unidades cadastradas."
    create_url_name = "dashboard_unidade_create"
    empty_message = "Nenhuma unidade cadastrada."
    columns = [
        {"label": "Nome", "field": "nome"},
    ]


class UnidadeCreateView(DashboardBaseCreateView):
    model = Unidade
    form_class = UnidadeForm
    success_url = reverse_lazy("dashboard_unidade_list")
    page_title_create = "Nova unidade"
    page_title_update = "Editar unidade"
    page_subtitle = "Preencha os dados da unidade."
    cancel_url_name = "dashboard_unidade_list"
    success_message = "Unidade cadastrada com sucesso."


class UnidadeUpdateView(DashboardBaseUpdateView):
    model = Unidade
    form_class = UnidadeForm
    success_url = reverse_lazy("dashboard_unidade_list")
    page_title_create = "Nova unidade"
    page_title_update = "Editar unidade"
    page_subtitle = "Preencha os dados da unidade."
    cancel_url_name = "dashboard_unidade_list"
    success_message = "Unidade atualizada com sucesso."


class UnidadeDeleteView(DashboardBaseDeleteView):
    model = Unidade
    success_url = reverse_lazy("dashboard_unidade_list")
    success_message = "Unidade excluída com sucesso."


# -------------------------
# NÍVEIS DE AVARIA
# -------------------------
class NivelAvariaListView(DashboardBaseListView):
    model = NivelAvaria
    ordering = ["nome"]
    page_title = "Níveis de avaria"
    page_subtitle = "Gerencie os níveis de avaria cadastrados."
    create_url_name = "dashboard_nivel_avaria_create"
    empty_message = "Nenhum nível de avaria cadastrado."
    columns = [
        {"label": "Nome", "field": "nome"},
    ]


class NivelAvariaCreateView(DashboardBaseCreateView):
    model = NivelAvaria
    form_class = NivelAvariaForm
    success_url = reverse_lazy("dashboard_nivel_avaria_list")
    page_title_create = "Novo nível de avaria"
    page_title_update = "Editar nível de avaria"
    page_subtitle = "Preencha os dados do nível de avaria."
    cancel_url_name = "dashboard_nivel_avaria_list"
    success_message = "Nível de avaria cadastrado com sucesso."


class NivelAvariaUpdateView(DashboardBaseUpdateView):
    model = NivelAvaria
    form_class = NivelAvariaForm
    success_url = reverse_lazy("dashboard_nivel_avaria_list")
    page_title_create = "Novo nível de avaria"
    page_title_update = "Editar nível de avaria"
    page_subtitle = "Preencha os dados do nível de avaria."
    cancel_url_name = "dashboard_nivel_avaria_list"
    success_message = "Nível de avaria atualizado com sucesso."


class NivelAvariaDeleteView(DashboardBaseDeleteView):
    model = NivelAvaria
    success_url = reverse_lazy("dashboard_nivel_avaria_list")
    success_message = "Nível de avaria excluído com sucesso."



# -------------------------
# VENDAS
# -------------------------

class VendaListView(DashboardPermissionMixin, ListView):
    model = Venda
    template_name = "dashboard/vendas/list.html"
    context_object_name = "vendas"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            Venda.objects
            .select_related("usuario", "forma_pagamento")
            .prefetch_related("itens")
            .order_by("-criado_em")
        )

        status = self.request.GET.get("status")
        mes = self.request.GET.get("mes")
        ano = self.request.GET.get("ano")

        if status:
            queryset = queryset.filter(status=status)

        if mes:
            queryset = queryset.filter(criado_em__month=mes)

        if ano:
            queryset = queryset.filter(criado_em__year=ano)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ano_atual = timezone.now().year
        context["status_atual"] = self.request.GET.get("status", "")
        context["mes_atual"] = self.request.GET.get("mes", "")
        context["ano_atual"] = self.request.GET.get("ano", "")
        context["status_choices"] = Venda._meta.get_field("status").choices
        context["meses"] = [
            (1, "Janeiro"),
            (2, "Fevereiro"),
            (3, "Março"),
            (4, "Abril"),
            (5, "Maio"),
            (6, "Junho"),
            (7, "Julho"),
            (8, "Agosto"),
            (9, "Setembro"),
            (10, "Outubro"),
            (11, "Novembro"),
            (12, "Dezembro"),
        ]
        context["anos"] = range(ano_atual - 1, ano_atual + 1)
        return context


class VendaDetailView(DashboardPermissionMixin, DetailView):
    model = Venda
    template_name = "dashboard/vendas/detail.html"
    context_object_name = "venda"

    def get_queryset(self):
        return (
            Venda.objects
            .select_related("usuario", "forma_pagamento")
            .prefetch_related("itens__produto", "itens__variacao")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = VendaStatusForm(instance=self.object)

        comprovante = None
        if self.object.forma_pagamento and self.object.forma_pagamento.codigo == "PIX":
            comprovante = self.object.comprovante_pix
        elif self.object.forma_pagamento and self.object.forma_pagamento.codigo == "VALE":
            comprovante = self.object.comprovante_vale

        comprovante_url = comprovante.url if comprovante else ""
        comprovante_nome = str(comprovante).lower() if comprovante else ""

        context["comprovante"] = comprovante
        context["comprovante_url"] = comprovante_url
        context["comprovante_is_pdf"] = comprovante_nome.endswith(".pdf")
        context["comprovante_is_image"] = comprovante_nome.endswith((".png", ".jpg", ".jpeg", ".webp"))
        return context


class VendaUpdateStatusView(DashboardPermissionMixin, UpdateView):
    model = Venda
    form_class = VendaStatusForm
    template_name = "dashboard/vendas/detail.html"
    context_object_name = "venda"

    def get_success_url(self):
        return reverse_lazy("dashboard_venda_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        status_anterior = Venda.objects.get(pk=form.instance.pk).status

        self.object = form.save()
        novo_status = self.object.status

        if (
            status_anterior != novo_status
            and novo_status in [Venda.Status.APROVADA, Venda.Status.CANCELADA, Venda.Status.CONCLUIDA]
        ):
            transaction.on_commit(
                lambda: enviar_email_status_venda_cliente(self.object)
            )

        messages.success(self.request, "Status da venda atualizado com sucesso.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Não foi possível atualizar a venda. Verifique os campos abaixo.")
        return self.render_to_response(self.get_context_data(form=form))


class VendaManageView(DashboardPermissionMixin, TemplateView):
    template_name = "dashboard/vendas/form.html"

    def dispatch(self, request, *args, **kwargs):
        self.object = None
        pk = kwargs.get("pk")
        if pk:
            self.object = Venda.objects.filter(pk=pk).first()
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, data=None, files=None):
        return VendaForm(data=data, files=files, instance=self.object)

    def get_formset(self, data=None, files=None):
        return VendaItemFormSet(
            data=data,
            files=files,
            instance=self.object,
            prefix="itens",
        )

    def get_produtos_info(self):
        produtos = (
            Produto.objects
            .filter(ativo=True)
            .prefetch_related("variacoes")
            .order_by("nome")
        )

        payload = {}
        for produto in produtos:
            variacoes = []
            if produto.usa_variacoes:
                variacoes = [
                    {
                        "id": variacao.id,
                        "label": " - ".join(
                            [
                                p for p in [
                                    variacao.get_categoria_display() if variacao.categoria else "",
                                    variacao.get_genero_display() if variacao.genero else "",
                                    variacao.get_faixa_etaria_display() if variacao.faixa_etaria else "",
                                    f"Tamanho {variacao.tamanho}" if variacao.tamanho else "",
                                    variacao.cor or "",
                                ] if p
                            ]
                        ) or f"Variação #{variacao.id}",
                        "estoque": int(variacao.quantidade or 0),
                    }
                    for variacao in produto.variacoes.filter(ativo=True).order_by(
                        "categoria", "genero", "faixa_etaria", "tamanho", "cor", "id"
                    )
                ]

            payload[str(produto.id)] = {
                "nome": produto.nome,
                "preco": str(produto.valor_venda),
                "estoque": int(produto.estoque_disponivel),
                "usa_variacoes": bool(produto.usa_variacoes),
                "variacoes": variacoes,
            }

        return payload

    def calcular_total_formset(self, formset):
        total = Decimal("0.00")

        for form in formset.forms:
            if not hasattr(form, "cleaned_data"):
                continue

            if form.cleaned_data.get("DELETE"):
                continue

            preco_unitario = form.cleaned_data.get("preco_unitario")
            quantidade = form.cleaned_data.get("quantidade") or 0

            if preco_unitario and quantidade:
                total += Decimal(preco_unitario) * Decimal(quantidade)

        return total.quantize(Decimal("0.01"))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        form = kwargs.get("form") or self.get_form()
        formset = kwargs.get("formset") or self.get_formset()

        total_venda = Decimal("0.00")
        if formset.is_bound and formset.is_valid():
            total_venda = self.calcular_total_formset(formset)
        elif self.object:
            total_venda = self.object.total or Decimal("0.00")

        context.update(
            {
                "object": self.object,
                "form": form,
                "formset": formset,
                "total_venda": total_venda,
                "produtos_info_json": json.dumps(self.get_produtos_info(), cls=DjangoJSONEncoder),
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())

    def post(self, request, *args, **kwargs):
        form = self.get_form(data=request.POST, files=request.FILES)
        formset = self.get_formset(data=request.POST, files=request.FILES)

        if not form.is_valid() or not formset.is_valid():
            messages.error(request, "Não foi possível salvar a venda. Verifique os campos abaixo.")
            return self.render_to_response(self.get_context_data(form=form, formset=formset))

        with transaction.atomic():
            venda = form.save(commit=False)
            is_create = venda.pk is None

            venda.total = Decimal("0.00")
            venda.save()

            formset.instance = venda
            itens = formset.save(commit=False)

            itens_existentes_ids = set(
                venda.itens.values_list("id", flat=True)
            ) if not is_create else set()

            itens_enviados_ids = set()

            for item in itens:
                produto = item.produto
                variacao = item.variacao
                quantidade = int(item.quantidade or 0)

                if produto.usa_variacoes:
                    if not variacao:
                        raise ValidationError("Produto com variação exige uma variação.")
                    if variacao.produto_id != produto.id:
                        raise ValidationError("A variação informada não pertence ao produto.")

                    if item.pk is None:
                        if quantidade > int(variacao.quantidade or 0):
                            raise ValidationError(
                                f"Estoque insuficiente para a variação '{variacao}'."
                            )
                        variacao.quantidade -= quantidade
                        variacao.save(update_fields=["quantidade"])
                    else:
                        antigo = VendaItem.objects.get(pk=item.pk)
                        delta = quantidade - int(antigo.quantidade or 0)

                        if antigo.variacao_id == variacao.id:
                            if delta > 0 and delta > int(variacao.quantidade or 0):
                                raise ValidationError(
                                    f"Estoque insuficiente para a variação '{variacao}'."
                                )
                            variacao.quantidade -= delta
                            variacao.save(update_fields=["quantidade"])
                        else:
                            if antigo.variacao_id:
                                antiga_variacao = antigo.variacao
                                antiga_variacao.quantidade += int(antigo.quantidade or 0)
                                antiga_variacao.save(update_fields=["quantidade"])

                            if quantidade > int(variacao.quantidade or 0):
                                raise ValidationError(
                                    f"Estoque insuficiente para a variação '{variacao}'."
                                )
                            variacao.quantidade -= quantidade
                            variacao.save(update_fields=["quantidade"])
                else:
                    if variacao:
                        raise ValidationError("Produto simples não pode receber variação.")

                    if item.pk is None:
                        if quantidade > int(produto.quantidade or 0):
                            raise ValidationError(
                                f"Estoque insuficiente para o produto '{produto.nome}'."
                            )
                        produto.quantidade -= quantidade
                        produto.save(update_fields=["quantidade"])
                    else:
                        antigo = VendaItem.objects.get(pk=item.pk)
                        delta = quantidade - int(antigo.quantidade or 0)

                        if delta > 0 and delta > int(produto.quantidade or 0):
                            raise ValidationError(
                                f"Estoque insuficiente para o produto '{produto.nome}'."
                            )

                        produto.quantidade -= delta
                        produto.save(update_fields=["quantidade"])

                item.venda = venda
                item.preco_unitario = produto.valor_venda
                item.save()
                itens_enviados_ids.add(item.id)

            # excluídos
            ids_para_excluir = list(itens_existentes_ids - itens_enviados_ids)
            if ids_para_excluir:
                for antigo in VendaItem.objects.filter(id__in=ids_para_excluir).select_related("produto", "variacao"):
                    if antigo.variacao_id:
                        variacao = antigo.variacao
                        variacao.quantidade += int(antigo.quantidade or 0)
                        variacao.save(update_fields=["quantidade"])
                    else:
                        produto = antigo.produto
                        produto.quantidade += int(antigo.quantidade or 0)
                        produto.save(update_fields=["quantidade"])

                VendaItem.objects.filter(id__in=ids_para_excluir).delete()

            for obj in formset.deleted_objects:
                if obj.pk and obj.pk not in ids_para_excluir:
                    if obj.variacao_id:
                        variacao = obj.variacao
                        variacao.quantidade += int(obj.quantidade or 0)
                        variacao.save(update_fields=["quantidade"])
                    else:
                        produto = obj.produto
                        produto.quantidade += int(obj.quantidade or 0)
                        produto.save(update_fields=["quantidade"])
                    obj.delete()

            venda.recalcular_total()

        messages.success(request, "Venda salva com sucesso.")
        return redirect("dashboard_venda_detail", pk=venda.pk)


class VendaExportXlsxView(DashboardPermissionMixin, View):
    def get_queryset(self):
        queryset = (
            Venda.objects
            .select_related("usuario", "forma_pagamento")
            .prefetch_related("itens__produto", "itens__variacao")
            .order_by("-criado_em")
        )

        status = (self.request.GET.get("status") or "").strip()
        mes = (self.request.GET.get("mes") or "").strip()
        ano = (self.request.GET.get("ano") or "").strip()

        if status:
            queryset = queryset.filter(status=status)

        if mes.isdigit():
            queryset = queryset.filter(criado_em__month=int(mes))

        if ano.isdigit():
            queryset = queryset.filter(criado_em__year=int(ano))

        return queryset

    def _get_usuario_nome(self, usuario):
        nome_completo = getattr(usuario, "get_full_name", lambda: "")()
        if nome_completo:
            return nome_completo

        return (
            getattr(usuario, "nome", "")
            or getattr(usuario, "username", "")
            or getattr(usuario, "email", "")
            or str(usuario)
        )

    def _get_usuario_cpf(self, usuario):
        return (
            getattr(usuario, "cpf", "")
            or getattr(usuario, "username", "")
            or ""
        )

    def _formatar_variacao(self, item):
        if not item.variacao:
            return ""

        partes = []

        if item.variacao.tamanho:
            partes.append(f"Tamanho {item.variacao.tamanho}")

        if item.variacao.genero:
            partes.append(item.variacao.get_genero_display())

        if item.variacao.faixa_etaria:
            partes.append(item.variacao.get_faixa_etaria_display())

        if getattr(item.variacao, "cor", None):
            partes.append(item.variacao.cor)

        return " / ".join(partes)

    def _formatar_item(self, item):
        descricao = item.produto.nome

        variacao_str = self._formatar_variacao(item)
        if variacao_str:
            descricao += f" - {variacao_str}"

        descricao += (
            f" (qtd: {item.quantidade}, "
            f"unit: R$ {Decimal(item.preco_unitario or 0):.2f}, "
            f"subtotal: R$ {Decimal(item.subtotal or 0):.2f})"
        )

        return descricao

    def get(self, request, *args, **kwargs):
        vendas = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas"

        headers = [
            "ID Venda",
            "Data",
            "Comprador",
            "CPF",
            "E-mail",
            "Forma de Pagamento",
            "Parcelas",
            "Status",
            "Valor Total",
            "Observação",
            "Itens",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for venda in vendas:
            usuario = venda.usuario

            itens_str = "\n".join(
                [self._formatar_item(item) for item in venda.itens.all()]
            )

            ws.append([
                venda.id,
                venda.criado_em.strftime("%d/%m/%Y %H:%M"),
                self._get_usuario_nome(usuario),
                self._get_usuario_cpf(usuario),
                getattr(usuario, "email", "") or "",
                str(venda.forma_pagamento),
                venda.parcelas,
                venda.get_status_display(),
                float(venda.total or Decimal("0.00")),
                venda.observacao or "",
                itens_str,
            ])

        widths = {
            "A": 12,
            "B": 20,
            "C": 30,
            "D": 20,
            "E": 30,
            "F": 20,
            "G": 12,
            "H": 15,
            "I": 15,
            "J": 40,
            "K": 100,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in range(2, ws.max_row + 1):
            ws[f"I{row}"].number_format = 'R$ #,##0.00'
            ws[f"B{row}"].alignment = Alignment(vertical="top")
            ws[f"C{row}"].alignment = Alignment(vertical="top")
            ws[f"D{row}"].alignment = Alignment(vertical="top")
            ws[f"E{row}"].alignment = Alignment(vertical="top")
            ws[f"F{row}"].alignment = Alignment(vertical="top")
            ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="top")
            ws[f"I{row}"].alignment = Alignment(horizontal="right", vertical="top")
            ws[f"J{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f"K{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"

        mes = (request.GET.get("mes") or "todos").strip()
        ano = (request.GET.get("ano") or "todos").strip()
        filename = f"vendas_{mes}_{ano}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

# -------------------------
# IMAGENS DE PRODUTOS
# -------------------------

class ProdutoImagemListView(DashboardPermissionMixin, ListView):
    model = ProdutoImagem
    template_name = "dashboard/produto-imagens/list.html"
    context_object_name = "imagens"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            ProdutoImagem.objects
            .select_related("produto")
            .order_by("produto__nome", "ordem", "id")
        )

        produto_id = self.request.GET.get("produto")
        if produto_id:
            queryset = queryset.filter(produto_id=produto_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["produtos"] = Produto.objects.order_by("nome")
        context["produto_atual"] = self.request.GET.get("produto", "")

        produto_id = self.request.GET.get("produto")
        context["produto_selecionado"] = None
        if produto_id:
            context["produto_selecionado"] = Produto.objects.filter(pk=produto_id).first()

        return context

class ProdutoImagemCreateView(DashboardPermissionMixin, CreateView):
    model = ProdutoImagem
    form_class = ProdutoImagemForm
    template_name = "dashboard/produto-imagens/form.html"
    success_url = reverse_lazy("dashboard_produto_imagem_list")

    def get_initial(self):
        initial = super().get_initial()
        produto_id = self.request.GET.get("produto")
        if produto_id:
            initial["produto"] = produto_id
        return initial

    def get_success_url(self):
        if self.object and self.object.produto_id:
            return f"{reverse_lazy('dashboard_produto_imagem_list')}?produto={self.object.produto_id}"
        return reverse_lazy("dashboard_produto_imagem_list")

    def form_valid(self, form):
        messages.success(self.request, "Imagem cadastrada com sucesso.")
        return super().form_valid(form)


class ProdutoImagemUpdateView(DashboardPermissionMixin, UpdateView):
    model = ProdutoImagem
    form_class = ProdutoImagemForm
    template_name = "dashboard/produto-imagens/form.html"
    success_url = reverse_lazy("dashboard_produto_imagem_list")

    def form_valid(self, form):
        messages.success(self.request, "Imagem atualizada com sucesso.")
        return super().form_valid(form)


class ProdutoImagemDeleteView(DashboardPermissionMixin, DeleteView):
    model = ProdutoImagem
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("dashboard_produto_imagem_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Imagem excluída com sucesso.")
        return super().delete(request, *args, **kwargs)
    
# -------------------------
# FORMAS DE PAGAMENTO
# -------------------------
class FormaPagamentoListView(DashboardPermissionMixin, ListView):
    model = FormaPagamento
    template_name = "dashboard/formas-pagamento/list.html"
    context_object_name = "formas_pagamento"
    paginate_by = 20
    ordering = ["codigo"]


class FormaPagamentoCreateView(DashboardPermissionMixin, CreateView):
    model = FormaPagamento
    form_class = FormaPagamentoForm
    template_name = "dashboard/formas-pagamento/form.html"
    success_url = reverse_lazy("dashboard_forma_pagamento_list")

    def form_valid(self, form):
        messages.success(self.request, "Forma de pagamento cadastrada com sucesso.")
        return super().form_valid(form)


class FormaPagamentoUpdateView(DashboardPermissionMixin, UpdateView):
    model = FormaPagamento
    form_class = FormaPagamentoForm
    template_name = "dashboard/formas-pagamento/form.html"
    success_url = reverse_lazy("dashboard_forma_pagamento_list")

    def form_valid(self, form):
        messages.success(self.request, "Forma de pagamento atualizada com sucesso.")
        return super().form_valid(form)


class FormaPagamentoDeleteView(DashboardPermissionMixin, DeleteView):
    model = FormaPagamento
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("dashboard_forma_pagamento_list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        if self.object.venda_set.exists():
            self.object.ativa = False
            self.object.save(update_fields=["ativa"])
            messages.warning(
                request,
                "Forma de pagamento já utilizada em vendas. Ela foi inativada em vez de excluída."
            )
        else:
            self.object.delete()
            messages.success(request, "Forma de pagamento excluída com sucesso.")

        return redirect(self.success_url)
    
# -------------------------
# REGRAS DE PARCELAMENTO DO VALE
# -------------------------
class RegraParcelamentoValeListView(DashboardBaseListView):
    model = RegraParcelamentoVale
    ordering = ["minimo"]
    page_title = "Regras de parcelamento"
    page_subtitle = "Gerencie as regras de parcelamento do vale."
    create_url_name = "dashboard_regra_vale_create"
    empty_message = "Nenhuma regra cadastrada."
    columns = [
        {"label": "Mínimo", "field": "minimo"},
        {"label": "Máximo", "field": "maximo"},
        {"label": "Parcelas", "field": "max_parcelas"},
    ]


class RegraParcelamentoValeCreateView(DashboardBaseCreateView):
    model = RegraParcelamentoVale
    form_class = RegraParcelamentoValeForm
    success_url = reverse_lazy("dashboard_regra_vale_list")
    page_title_create = "Nova regra de parcelamento"
    page_title_update = "Editar regra de parcelamento"
    page_subtitle = "Preencha os dados da regra."
    cancel_url_name = "dashboard_regra_vale_list"
    success_message = "Regra de parcelamento cadastrada com sucesso."


class RegraParcelamentoValeUpdateView(DashboardBaseUpdateView):
    model = RegraParcelamentoVale
    form_class = RegraParcelamentoValeForm
    success_url = reverse_lazy("dashboard_regra_vale_list")
    page_title_create = "Nova regra de parcelamento"
    page_title_update = "Editar regra de parcelamento"
    page_subtitle = "Preencha os dados da regra."
    cancel_url_name = "dashboard_regra_vale_list"
    success_message = "Regra de parcelamento atualizada com sucesso."


class RegraParcelamentoValeDeleteView(DashboardBaseDeleteView):
    model = RegraParcelamentoVale
    success_url = reverse_lazy("dashboard_regra_vale_list")
    success_message = "Regra de parcelamento excluída com sucesso."